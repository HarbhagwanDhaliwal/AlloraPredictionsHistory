import asyncio
import subprocess
import yaml
import math
from config import RPC_URLS, SLEEP_TIME, ZERO_SLEEP_TIME, ALLORA_CHAIN_BLOCK_TIME_IN_SECONDS
import time
from datetime import datetime
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os
import random


def get_random_rpc():
    return random.choice(RPC_URLS)


async def run_command(command):
    try:
        process = await asyncio.create_subprocess_shell(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE
        )
        stdout, stderr = await asyncio.wait_for(process.communicate(), timeout=15)

        if process.returncode == 0:
            return stdout.decode().strip()
        else:
            raise Exception(f"Command failed: {stderr.decode().strip()}")
    except asyncio.TimeoutError:
        raise Exception("Command timed out")


async def execute_with_retry(command, retries=3):
    for attempt in range(retries):
        try:
            output = await run_command(command)
            return output
        except Exception as e:
            error_message = str(e)
            if "lowest height is" in error_message and "is not available" in error_message:
                print("❌ Block height is too old and has been pruned. Cannot fetch data.")
                return False
            print(f"Attempt {attempt + 1} failed: {e}")
            if attempt + 1 < retries:
                print(f"Retrying after {SLEEP_TIME} seconds...")
                await asyncio.sleep(SLEEP_TIME)
            else:
                print(f"All attempts failed for command: {command} \nError: {e}")
                return False
    return False


async def get_latest_inference_block(topic_id: int) -> int:
    rpc_url = get_random_rpc()
    command = [
        "allorad",
        "q",
        "emissions",
        "latest-network-inferences",
        str(topic_id),
        "--node",
        rpc_url
    ]
    command_str = " ".join(command)

    try:
        output = await execute_with_retry(command_str)
        if not output:
            raise ValueError("Failed to get latest network inferences")

        for line in output.split('\n'):
            if line.strip().startswith("inference_block_height:"):
                block_height = line.split(':')[1].strip().strip('"')
                return int(block_height)

        raise ValueError("Could not find inference_block_height in command output")

    except Exception as e:
        print(f"Error getting latest inference block: {e}")
        raise


async def is_topic_valid(topic_id):
    rpc_url = get_random_rpc()
    command = f"allorad q emissions topic-exists {topic_id} --node {rpc_url}"
    try:
        response = await execute_with_retry(command)
        data = yaml.safe_load(response)
        return bool(data and 'exists' in data and data['exists'])
    except Exception as e:
        print(f"Validation failed: {e}")
        return False


async def get_topic(topic_id):
    if not await is_topic_valid(topic_id):
        return None, None, None

    rpc_url = get_random_rpc()
    command = f"allorad q emissions topic {topic_id} --node {rpc_url}"

    try:
        response = await execute_with_retry(command)
        data = yaml.safe_load(response)
        if data is None or 'topic' not in data:
            return None, None, None

        topic_data = data['topic']
        epoch_last_ended = topic_data.get('epoch_last_ended')
        epoch_length = topic_data.get('epoch_length')
        metadata = topic_data.get('metadata', '')

        return epoch_last_ended, epoch_length, metadata
    except Exception as e:
        print(f"Error getting topic info: {e}")
        return None, None, None


async def get_inference_for_block(topic_id, block_height):
    rpc_url = get_random_rpc()
    command = (
        f"allorad q emissions network-inferences-at-block-outlier-resistant "
        f"{topic_id} {block_height} --node {rpc_url}"
    )

    try:
        response = await execute_with_retry(command)
        data = yaml.safe_load(response)

        if data is None:
            print("No data received from command")
            return topic_id, None, None

        network_inferences = data.get('network_inferences', {})
        if not network_inferences:
            print("No network_inferences found in response")
            return topic_id, None, None

        combined_value = network_inferences.get('combined_value')
        if combined_value is None:
            print("No combined_value found in network_inferences")
            return topic_id, None, None

        return topic_id, block_height, combined_value
    except Exception as e:
        print(f"Failed to fetch inference: {e}")
        return topic_id, None, None


async def get_block_time(block_height):
    rpc_url = get_random_rpc()
    command = f"allorad query block --type=height {block_height} --node {rpc_url}"

    try:
        response = await execute_with_retry(command)
        data = yaml.safe_load(response)

        if not data or 'header' not in data:
            print("No header data in block response")
            return None

        header = data['header']
        block_time_str = header.get('time')

        if not block_time_str:
            print("No time field in block header")
            return None

        try:
            if '.' in block_time_str:
                block_time_str = block_time_str.split('.')[0] + 'Z'
            block_time = datetime.fromisoformat(block_time_str.replace('Z', '+00:00'))
            return block_time
        except ValueError as e:
            print(f"Error parsing block time: {e}")
            return None

    except Exception as e:
        print(f"Failed to fetch block time: {e}")
        return None


def calculate_blocks_from_timeframe(timeframe_str: str) -> int:
    if timeframe_str == "1Hour":
        seconds = 60 * 60
    elif timeframe_str == "1Day":
        seconds = 24 * 60 * 60
    elif timeframe_str == "1Week":
        seconds = 7 * 24 * 60 * 60
    elif timeframe_str == "1Month":
        seconds = 30 * 24 * 60 * 60
    else:
        raise ValueError("Invalid timeframe selected")

    return math.ceil(seconds / ALLORA_CHAIN_BLOCK_TIME_IN_SECONDS)


async def collect_inference_data(topic_id, from_block):
    epoch_last_ended, epoch_length, metadata = await get_topic(topic_id)

    if epoch_last_ended is None or epoch_length is None:
        print("Failed to get topic information.")
        return None

    print(f"\nTopic Metadata: {metadata}")
    print(f"Epoch Last Ended: {epoch_last_ended}")
    print(f"Epoch Length: {epoch_length}\n")

    current_block = int(epoch_last_ended)
    from_block = int(from_block)
    epoch_length = int(epoch_length)
    data = []
    consecutive_zero_count = 0
    MAX_CONSECUTIVE_ZEROS = 3

    while current_block >= from_block:
        print(f"Processing block {current_block}...")

        _, block, combined_value = await get_inference_for_block(topic_id, current_block)

        if block is None or combined_value is None:
            print(f"Failed to get inference data for block {current_block}")
            current_block -= epoch_length
            continue

        block_time = await get_block_time(current_block)
        if block_time is None:
            print(f"Failed to get timestamp for block {current_block}")
            current_block -= epoch_length
            continue

        if float(combined_value) == 0:
            consecutive_zero_count += 1
            print(f"Zero combined value at block {current_block}. Consecutive zeros: {consecutive_zero_count}")

            print(f"Sleeping for {ZERO_SLEEP_TIME} seconds due to zero value...")
            await asyncio.sleep(ZERO_SLEEP_TIME)

            if consecutive_zero_count >= MAX_CONSECUTIVE_ZEROS:
                print("Reached maximum consecutive zero values. Stopping collection.")
                break

            current_block -= epoch_length
            continue
        else:
            if consecutive_zero_count > 0:
                print("Non-zero value received. Resetting consecutive zero counter.")
            consecutive_zero_count = 0

        data.append({
            'block_height': current_block,
            'block_timestamp': block_time.isoformat(),
            'allora_predicted_value': combined_value,
        })

        print(f"Collected data for block {current_block}: {combined_value} at {block_time.isoformat()}")

        current_block -= epoch_length

    if data:
        df = pd.DataFrame(data)
        df.columns = [col.upper() for col in df.columns]

        os.makedirs("data", exist_ok=True)

        filename = f"data/allora_chain_data_topic_{topic_id}.xlsx"
        df.to_excel(filename, index=False)

        workbook = load_workbook(filename)
        sheet = workbook.active

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max_length + 2

        workbook.save(filename)

        print(f"\nSuccessfully saved data to {filename}")
        return filename
    else:
        print("\nNo valid data collected")
        return None
